import csv
from dataclasses import dataclass
from pathlib import Path

import numpy as np


@dataclass(frozen=True)
class ImportedSeries:
    label: str
    values: tuple[float, ...]


@dataclass(frozen=True)
class ImportedTable:
    x_label: str
    x_values: tuple[float, ...]
    series: tuple[ImportedSeries, ...]


def _read_csv_rows(path):
    last_error = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                return list(csv.reader(file))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError("CSV文件编码无法识别，请使用UTF-8或GB18030") from last_error


def _read_xlsx_rows(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("读取Excel文件需要安装openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def load_table(file_path):
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw_rows = _read_csv_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        raw_rows = _read_xlsx_rows(path)
    else:
        raise ValueError("仅支持CSV、XLSX或XLSM文件")
    rows = [row for row in raw_rows if any(cell is not None and str(cell).strip() for cell in row)]
    if len(rows) < 2 or len(rows[0]) < 2:
        raise ValueError("表格至少需要标题行、一个数据行和两列数据")

    headers = [str(cell).strip() or f"数据列{index + 1}" for index, cell in enumerate(rows[0])]
    column_count = len(headers)
    columns = [[] for _ in range(column_count)]
    for row_number, row in enumerate(rows[1:], start=2):
        if len(row) != column_count:
            raise ValueError(f"第{row_number}行的列数与标题行不一致")
        for column_index, cell in enumerate(row):
            try:
                value = float(str(cell).strip())
            except ValueError as exc:
                raise ValueError(f"第{row_number}行第{column_index + 1}列不是有效数字") from exc
            if not np.isfinite(value):
                raise ValueError(f"第{row_number}行第{column_index + 1}列必须是有限数字")
            columns[column_index].append(value)

    return ImportedTable(
        x_label=headers[0],
        x_values=tuple(columns[0]),
        series=tuple(
            ImportedSeries(label=headers[index], values=tuple(columns[index]))
            for index in range(1, column_count)))
